from typing import Any
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Saver(object):

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def save_power(self, power_dict, path):
        self.ws = self.wb.create_sheet("power")
        self.ws["A1"] = "waveLength"
        self.ws["B1"] = "value"
        for index, key in enumerate(power_dict.keys()):
            self.ws["A{}".format(str(index + 2))] = key
            self.ws["B{}".format(str(index + 2))] = power_dict[key]
        self.write_file(path)

    def write_file(self, path):
        self.wb.save(path)